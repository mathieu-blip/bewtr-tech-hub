# -*- coding: utf-8 -*-
import csv, json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCR="/tmp/claude-0/-home-user-bewtr-tech-hub/9d71d817-09e8-524b-89cc-8e99b6e4ed28/scratchpad"
OUT="/home/user/bewtr-tech-hub/out/BE-WTR_commandes-spare-parts-par-pays_2026-08-26.xlsx"
ref={r["ref"]:r for r in csv.DictReader(open("/home/user/bewtr-tech-hub/docs/audit-spare-parts-2022.csv",encoding="utf-8-sig"))}
B24=json.load(open(f"{SCR}/blupura2024.json"))
IDX={}
for r in B24:
    for k in (r["cur"],r["alt"],r["new"]):
        if k: IDX.setdefault(k,r)
def blu(sku_or_rf):
    """retourne (code2024, prix_eur, cond, desc_en) depuis la ref fournisseur"""
    m=IDX.get(sku_or_rf)
    if not m: return (None,None,1,None)
    mm=re.search(r'N\.\s*(\d+)\s*PCS', m["en"], re.I)
    return (m["new"], m["price"], int(mm.group(1)) if mm else 1, m["en"])
def pchf(s):
    try: return float(ref.get(s,{}).get("prix_catalogue",""))
    except: return None

A="Arial"
H1=Font(name=A,size=14,bold=True); H2=Font(name=A,size=11,bold=True)
BASE=Font(name=A,size=10); BOLD=Font(name=A,size=10,bold=True)
ITAL=Font(name=A,size=9,italic=True,color="595959"); WB_=Font(name=A,size=10,bold=True,color="FFFFFF")
BLUE=Font(name=A,size=10,color="0000FF")
HDR=PatternFill("solid",fgColor="1F3864"); GREY=PatternFill("solid",fgColor="BFBFBF")
WARN=PatternFill("solid",fgColor="FFF2CC"); TOT=PatternFill("solid",fgColor="E2EFDA")
t=Side(style="thin",color="BFBFBF"); BORD=Border(left=t,right=t,top=t,bottom=t)
MONEY='#,##0.00;(#,##0.00);-'; INT='#,##0;(#,##0);-'
def hdr(ws,row,labs,widths):
    for i,l in enumerate(labs,1):
        c=ws.cell(row=row,column=i,value=l); c.font=WB_; c.fill=HDR; c.border=BORD
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.row_dimensions[row].height=36
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
def titre(ws,txt,sub=None):
    ws["A1"]=txt; ws["A1"].font=H1
    if sub: ws["A2"]=sub; ws["A2"].font=ITAL

L=[]
def add(pays,sku,lib,qte,orig,item,note="",rf_over=None,four_over=None):
    L.append(dict(pays=pays,sku=sku,lib=lib,qte=qte,orig=orig,item=item,note=note,
                  rf_over=rf_over,four_over=four_over))
# FRANCE — commande
for s,q in [("BW-0416",5),("BW-0183",20),("BW-0184",6),("BW-0417",3),("BW-0429",30),("BW-0421",30),
            ("BW-0419",2),("BW-0420",2),("BW-0428",10),("BW-0186",20),("BW-0424",10),("BW-0425",5),
            ("BW-0426",5),("BW-0195",5)]:
    add("France",s,None,q,"Commande Monday","12889411743")
add("France",None,"Compresseur BOX 80",1,"Claim ouvert","12759710752",
    "Aucun compresseur au référentiel BOX 80. Tarif Blupura 2024 : 5 compresseurs possibles — modèle à identifier.")
add("France",None,"Électrovanne d'entrée BOX 80",1,"Claim ouvert","12759640381",
    "Aucune électrovanne d'entrée au référentiel BOX 80. Tarif 2024 : 4 candidats (150069, 750376) — à identifier.")
# SUISSE — commande
add("Suisse","BW-0191",None,30,"Commande Monday","12889827793",
    "Choix validé Blupura 750244. Le tarif 2024 prouve que 120232 (BW-0191) et 750244 sont le même article (760032). "
    "BW-0988 « Axial fan 120x120x25 » du BOX 20 est un doublon de SKU pour cette même pièce.",rf_over="750244")
add("Suisse","BW-0301",None,20,"Commande Monday","12889827793",
    "Demandé en 230-12V ; BW-0301 est en 230V/24VDC. Choix validé.")
add("Suisse","BW-0507",None,15,"Commande Monday","12889827793",
    "Ligne dupliquée dans Monday — 15 retenu. Pièce Borg&Overström, absente du tarif Blupura : prix à demander.")
add("Suisse","BW-0164",None,20,"Commande Monday","12889827793","Ligne dupliquée dans Monday — 20 retenu.")
add("Suisse",None,"Sonde BOX PRO2 V2",20,"Commande Monday","12889827793",
    "Aucune sonde PRO2 au référentiel — SKU à renseigner.")
# SUISSE — claims
add("Suisse","BW-0179",None,1,"Claim ouvert","12217762646","Diagnostic technicien : « il faut sûrement changer l'électrovanne ».")
add("Suisse","BW-0974",None,1,"Claim ouvert","12345567586","« Fuite au chapeau de la gearbox ». Superinox, conditionné par 5.")
add("Suisse","BW-0994",None,1,"Claim ouvert","12891544139","« Carrosserie HS ». Pièce BOX 20 sans référence fournisseur.")
add("Suisse",None,"Bouton BAR2 (position haut-droite)",1,"Claim ouvert","12498897081",
    "4 boutons BAR2 possibles (810695-810698) — position à préciser.")
add("Suisse",None,"Fuite CO2 BOX 20 — pièce à identifier",1,"Claim ouvert","12891695968",
    "Fuite long terme, pièce non diagnostiquée.")
# UK / EMIRATS
add("UK","BW-0295",None,1,"Claim ouvert","12232314055","« Drip tray cracked and leaking ».")
add("Émirats","BW-0172",None,1,"Claim ouvert","18083626539","« Release pressure valve — new spare part needed ».")
add("Émirats",None,"Spare parts manual capper (Tower)",1,"Claim ouvert","10943979984",
    "Machine Tower / capper hors référentiel — SKU à créer.")

wb=Workbook()
ws=wb.active; ws.title="Commandes par pays"
titre(ws,"Pièces à commander par pays — SKU, référence fournisseur et prix",
      "Monday « Technical troubleshooting » au 26.08.2026, croisé avec le tarif Blupura « Trascodifica REV00 » du 08.11.2024 "
      "et la base SharePoint de septembre 2022. Cellule SKU grisée = référence à renseigner.")
hdr(ws,4,["Pays","SKU BW","Désignation","Réf. fourn.","Code Blupura 2024","Fournisseur","Qté (pcs)",
          "Cond. (pcs/lot)","Unités à cder","Prix unit. EUR (2024)","Prix cat. CHF (2022)",
          "Total EUR (Blupura)","Total CHF (autres fourn.)","Origine","Item Monday","Remarque"],
    [10,10,36,11,14,14,9,10,11,15,15,11,11,15,13,50])
r=5; first=r
for x in L:
    s=x["sku"]; d=ref.get(s,{}) if s else {}
    lib=d.get("designation") or x["lib"]
    rf=x["rf_over"] or d.get("ref_fournisseur") or ""
    fo=x["four_over"] or d.get("fournisseur") or ""
    code,eur,cond,_=blu(rf) if rf else (None,None,1,None)
    chf=pchf(s) if s else None
    ws.cell(row=r,column=1,value=x["pays"]).font=BOLD
    c=ws.cell(row=r,column=2,value=s or ""); c.font=BOLD
    if not s: c.fill=GREY
    ws.cell(row=r,column=3,value=lib).font=BASE
    ws.cell(row=r,column=4,value=rf or "—").font=BASE
    ws.cell(row=r,column=5,value=code or "—").font=BASE
    ws.cell(row=r,column=6,value=fo or "—").font=BASE
    ws.cell(row=r,column=7,value=x["qte"]).font=BLUE
    ws.cell(row=r,column=8,value=cond).font=BASE
    ws.cell(row=r,column=9,value=f"=IF(H{r}<=1,G{r},ROUNDUP(G{r}/H{r},0))").font=BASE
    ws.cell(row=r,column=10,value=eur).font=BASE
    ws.cell(row=r,column=11,value=chf).font=BASE
    ws.cell(row=r,column=12,value=f'=IF(N(J{r})=0,"",I{r}*J{r})').font=BASE
    ws.cell(row=r,column=13,value=f'=IF(N(J{r})>0,"",IF(N(K{r})=0,"",G{r}*K{r}))').font=BASE
    ws.cell(row=r,column=14,value=x["orig"]).font=BASE
    ws.cell(row=r,column=15,value=x["item"]).font=BASE
    ws.cell(row=r,column=16,value=x["note"]).font=BASE
    for cc in range(1,17):
        cell=ws.cell(row=r,column=cc); cell.border=BORD
        cell.alignment=Alignment(wrap_text=(cc in(3,16)),vertical="top")
    for cc in(10,11,12,13): ws.cell(row=r,column=cc).number_format=MONEY
    for cc in(7,8,9): ws.cell(row=r,column=cc).number_format=INT
    if cond>1: ws.cell(row=r,column=8).fill=WARN
    if x["note"]: ws.cell(row=r,column=16).fill=WARN
    r+=1
last=r-1
ws.cell(row=r,column=1,value="TOTAL").font=BOLD
ws.cell(row=r,column=7,value=f"=SUM(G{first}:G{last})").font=BOLD
ws.cell(row=r,column=12,value=f"=SUM(L{first}:L{last})").font=BOLD
ws.cell(row=r,column=13,value=f"=SUM(M{first}:M{last})").font=BOLD
ws.cell(row=r,column=16,value=f'=COUNTBLANK(B{first}:B{last})&" SKU à renseigner · "&COUNTIFS(J{first}:J{last},"",K{first}:K{last},"")&" ligne(s) sans aucun prix"').font=BOLD
for cc in range(1,17):
    cell=ws.cell(row=r,column=cc); cell.border=BORD; cell.fill=TOT
ws.cell(row=r,column=7).number_format=INT
for cc in(12,13): ws.cell(row=r,column=cc).number_format=MONEY
GRAND=r
ws.freeze_panes="D5"; ws.auto_filter.ref=f"A4:P{last}"
r+=2
for txt in ["Cellule SKU grisée : pièce identifiée mais sans référence BE WTR — à créer ou à faire préciser.",
  "Colonne « Cond. » surlignée : Blupura vend cet article par lot. « Unités à cder » arrondit la quantité demandée au lot supérieur, "
  "et le total EUR porte sur les lots, pas sur les pièces.",
  "Prix unit. EUR = tarif d'achat Blupura du 08.11.2024 (le plus récent). Prix cat. CHF = prix catalogue BE WTR 2022, "
  "conservé pour les pièces Borg&Overström et Superinox absentes du tarif Blupura.",
  "Les deux totaux sont disjoints et s'additionnent : le total CHF ne porte que sur les lignes sans prix d'achat Blupura "
  "(Borg&Overström, Superinox). Une ligne chiffrée en euros n'est jamais recomptée en francs.",
  "Bleu = quantité issue de Monday. Noir = donnée de référentiel ou calcul."]:
    ws.cell(row=r,column=1,value=txt).font=ITAL; r+=1

# ---------- SYNTHESE ----------
ws=wb.create_sheet("Synthèse",0)
titre(ws,"Commandes de pièces détachées par pays","Extrait du board Monday « Technical troubleshooting » le 26.08.2026.")
hdr(ws,4,["Pays","Lignes","Pièces","dont commande","dont claim","SKU à renseigner","Lignes sans prix","Claims ouverts"],
    [14,9,10,15,12,17,17,15])
CL={"Suisse":38,"France":10,"Émirats":2,"UK":1,"Espagne":1,"Canada":1}
def has_price(x):
    s=x["sku"]; d=ref.get(s,{}) if s else {}
    rf=x["rf_over"] or d.get("ref_fournisseur") or ""
    return (blu(rf)[1] is not None) or (pchf(s) is not None if s else False)
r=5
for pays in ["France","Suisse","UK","Émirats","Espagne","Canada"]:
    li=[x for x in L if x["pays"]==pays]
    for i,v in enumerate([pays,len(li),sum(x["qte"] for x in li),
        sum(x["qte"] for x in li if x["orig"]=="Commande Monday"),
        sum(x["qte"] for x in li if x["orig"]=="Claim ouvert"),
        sum(1 for x in li if not x["sku"]),
        sum(1 for x in li if not has_price(x)), CL.get(pays,0)],1):
        c=ws.cell(row=r,column=i,value=v); c.font=BOLD if i==1 else BASE; c.border=BORD
        if i>1: c.number_format=INT
    r+=1
ws.cell(row=r,column=1,value="TOTAL").font=BOLD
for cc in range(2,9):
    Lt=get_column_letter(cc)
    ws.cell(row=r,column=cc,value=f"=SUM({Lt}5:{Lt}{r-1})").font=BOLD
    ws.cell(row=r,column=cc).number_format=INT
for cc in range(1,9):
    cell=ws.cell(row=r,column=cc); cell.border=BORD; cell.fill=TOT
r+=2
ws.cell(row=r,column=1,value="Montant Blupura (EUR)").font=H2
ws.cell(row=r,column=3,value=f"='Commandes par pays'!L{GRAND}").font=BOLD
ws.cell(row=r,column=3).number_format=MONEY; r+=1
ws.cell(row=r,column=1,value="Montant Borg&Overström (CHF)").font=H2
ws.cell(row=r,column=3,value=f"='Commandes par pays'!M{GRAND}").font=BOLD
ws.cell(row=r,column=3).number_format=MONEY; r+=2
for txt in [
 "Le tarif Blupura « Trascodifica » du 08.11.2024 change la donne : 12 des 14 lignes françaises sont désormais chiffrables, "
 "contre 4 avec la seule base de 2022. Seuls le support de pompe immersion (110264) et le kit ice bank overflow (750019) "
 "restent absents du tarif et doivent être devisés.",
 "Ce tarif révèle aussi que les prix catalogue 2022 sont environ le double du prix d'achat réel : "
 "pompe immersion 21 CHF contre 11,55 EUR, pompe 100 L/h 121 CHF contre 75,60 EUR, ventilateur 40 CHF contre 23,80 EUR.",
 "Attention aux conditionnements : le pied PVC (140118) se vend par 4 et le passe-cloison 8 mm (140099) par 10. "
 "Les 30 pieds et 10 passe-cloisons demandés correspondent donc à 8 et 1 lots.",
 "Espagne et Canada n'ont aucune pièce à commander : claim espagnol vide, claim canadien en attente d'un frigoriste local.",
 "France : 4 moteurs de pompe CO2 sont en panne pour 3 commandés, alors que 20 thermostats sont commandés pour 3 pannes.",
]:
    c=ws.cell(row=r,column=1,value=txt); c.font=BASE
    c.alignment=Alignment(wrap_text=True,vertical="top")
    ws.merge_cells(start_row=r,start_column=1,end_row=r+1,end_column=8); r+=3

# ---------- VERIFICATION CLAIMS ----------
ws=wb.create_sheet("Vérification claims")
titre(ws,"Vérification des 53 lignes ouvertes du Monday",
      "Statuts non clôturés au 26.08.2026 : New (to repair) 47, In progress 3, At supplier 2, Stucked 1.")
hdr(ws,4,["Item","Pays","Produit","Date","Statut","Sujet","Description","Verdict","Pièce retenue"],
    [13,11,26,11,16,32,52,26,32])
op=json.load(open(f"{SCR}/open53.json"))
V={"12889411743":("Commande formalisée","14 SKU BOX 80"),"12889827793":("Commande formalisée","5 lignes"),
 "12759710752":("Pièce déduite","Compresseur — SKU à créer"),"12759640381":("Pièce déduite","Électrovanne entrée — SKU à créer"),
 "12217762646":("Pièce déduite","BW-0179 PRO2 Solenoids"),"12345567586":("Pièce déduite","BW-0974 Gearbox - Hat"),
 "12891544139":("Pièce déduite","BW-0994 Outer casing BOX 20"),"12498897081":("Pièce déduite","Bouton BAR2 — à préciser"),
 "12891695968":("Pièce déduite","Fuite CO2 — à diagnostiquer"),"12232314055":("Pièce déduite","BW-0295 BAR2 Drip tray"),
 "18083626539":("Pièce déduite","BW-0172 8mm PRV"),"10943979984":("Pièce déduite","Capper Tower — SKU à créer"),
 "12890647974":("Couvert par la commande","Ventilateur BW-0191 (30 pcs)"),
 "12189511251":("Couvert par la commande","Ventilateur BW-0191 (30 pcs)"),
 "12891290504":("Exclu sur décision","Panneau tactile PRO2"),"12488288189":("Exclu sur décision","Panneau tactile PRO2"),
 "12488429466":("Sans pièce","Machine échangée"),"9340197921":("Sans pièce","Attente frigoriste local"),
 "12890697274":("Sans pièce","Claim vide")}
PF={"Switzerland":"Suisse","France":"France","UK":"UK","United Arab Emirates":"Émirats","Spain":"Espagne","Canada":"Canada"}
FS={"12734652195":"Moteur pompe CO2 + thermostat","12759584213":"Ventilateur + froid","12759667201":"Moteur pompe CO2",
 "12759598322":"Ventilateur + froid","12759621953":"Moteur pompe CO2 + ventilateur","12759641294":"Ventilateur + thermostat",
 "12759687821":"Thermostat"}
ORD={"Commande formalisée":0,"Pièce déduite":1,"Couvert par la commande":2,"Exclu sur décision":3,"Sans pièce":4,"Diagnostic à faire":5}
rr=[]
for x in op:
    v,p=V.get(x["id"],(None,None))
    if v is None:
        v,p=("Couvert par la commande",FS[x["id"]]) if x["id"] in FS else ("Diagnostic à faire","—")
    rr.append((x,v,p))
rr.sort(key=lambda z:(ORD[z[1]],PF.get(z[0]["pays"],""),z[0]["prod"] or ""))
r=5
for x,v,p in rr:
    for i,val in enumerate([x["id"],PF.get(x["pays"],x["pays"]),x["prod"] or "—",x["date"] or "—",
                            x["statut"],x["name"][:60],(x["desc"] or "")[:150],v,p],1):
        c=ws.cell(row=r,column=i,value=val); c.font=BOLD if i in(2,8) else BASE; c.border=BORD
        c.alignment=Alignment(wrap_text=(i in(6,7,9)),vertical="top")
    if v=="Diagnostic à faire": ws.cell(row=r,column=8).fill=WARN
    r+=1
ws.freeze_panes="A5"; ws.auto_filter.ref=f"A4:I{r-1}"
r+=1
ws.cell(row=r,column=1,value="« Diagnostic à faire » = la description ne permet pas d'identifier une pièce "
        "(« Box not working correctly », « Box gelée », « Subject »). Ces lignes ne génèrent pas de commande en l'état.").font=ITAL

# ---------- TARIF BLUPURA 2024 ----------
ws=wb.create_sheet("Tarif Blupura 11-2024")
titre(ws,"Tarif Blupura — Trascodifica REV00, 08.11.2024","545 références. Code actuel / alternatif → nouveau code de substitution, prix net unitaire en euros.")
hdr(ws,4,["#","Code actuel","Code alternatif","Nouveau code","Désignation (EN)","Prix net (EUR)","Cond. (pcs/lot)","SKU BW correspondant"],
    [6,13,15,13,62,13,13,18])
RF2SKU={}
for s,d in ref.items():
    if d["ref_fournisseur"]: RF2SKU.setdefault(d["ref_fournisseur"],s)
r=5
for m in B24:
    mm=re.search(r'N\.\s*(\d+)\s*PCS',m["en"],re.I)
    cond=int(mm.group(1)) if mm else 1
    sku=RF2SKU.get(m["cur"]) or RF2SKU.get(m["alt"]) or ""
    for i,v in enumerate([m["item"],m["cur"] or "—",m["alt"] or "—",m["new"],m["en"],m["price"],cond,sku or "—"],1):
        c=ws.cell(row=r,column=i,value=v); c.font=BOLD if i==8 and sku else BASE; c.border=BORD
    ws.cell(row=r,column=6).number_format=MONEY
    if sku: ws.cell(row=r,column=8).fill=TOT
    r+=1
ws.freeze_panes="A5"; ws.auto_filter.ref=f"A4:H{r-1}"

# ---------- REFERENTIEL ----------
ws=wb.create_sheet("Référentiel SKU & prix")
titre(ws,"Référentiel pièces détachées — 294 SKU",
      "Hub technicien croisé avec la base SharePoint 2022 et le tarif Blupura 08.11.2024.")
hdr(ws,4,["SKU BW","Désignation","Machines (hub)","Fournisseur","Réf. fourn.","Code Blupura 2024",
          "Prix achat EUR (2024)","Prix cat. CHF (2022)","Remarque"],[11,38,32,15,12,14,16,16,26])
r=5
for d in sorted(ref.values(),key=lambda x:x["ref"]):
    s=d["ref"]; rf=d["ref_fournisseur"]; code,eur,cond,_=blu(rf) if rf else (None,None,1,None)
    raw=d["prix_catalogue"]
    rem=("Chiffrable via tarif 2024" if eur is not None and raw in("AS*","") else
         ("Prix à demander (AS*)" if raw=="AS*" else ("Absente de la base 2022" if raw=="" and not d["fournisseur"] else "")))
    for i,v in enumerate([s,d["designation"],d["machines_hub"],d["fournisseur"],rf,code or "—",eur,pchf(s),rem],1):
        c=ws.cell(row=r,column=i,value=v); c.font=BOLD if i==1 else BASE; c.border=BORD
    for cc in(7,8): ws.cell(row=r,column=cc).number_format=MONEY
    if rem.startswith("Chiffrable"): ws.cell(row=r,column=9).fill=TOT
    elif rem: ws.cell(row=r,column=9).fill=WARN
    r+=1
ws.freeze_panes="A5"; ws.auto_filter.ref=f"A4:I{r-1}"

for s in wb.worksheets:
    if s.title in("Synthèse",): s.sheet_view.showGridLines=False
wb.save(OUT); print("OK",OUT)
