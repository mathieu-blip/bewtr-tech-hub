# -*- coding: utf-8 -*-
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCR = "/tmp/claude-0/-home-user-bewtr-tech-hub/9d71d817-09e8-524b-89cc-8e99b6e4ed28/scratchpad"
CSVP = "/home/user/bewtr-tech-hub/docs/audit-spare-parts-2022.csv"
OUT = "/home/user/bewtr-tech-hub/out/BE-WTR_commandes-spare-parts_2026-08-26.xlsx"

# ---------- referentiel ----------
ref = {}
rows_ref = []
with open(CSVP, encoding="utf-8-sig") as f:
    for r in csv.DictReader(f):
        ref[r["ref"]] = r
        rows_ref.append(r)

# Blupura purchase price list (EUR) — "Suggested Spare Parts list for Service - Price List 2022 - BE WTR.xlsx"
BLUPURA_EUR = {
 "750244":22.50,"750243":21.60,"150068":2.70,"750230":27.30,"130009":72.00,
 "810320":10.50,"130053":138.70,"130049":68.60,"140092":15.40,"140194":15.40,
 "150031":12.20,"150391":11.00,"150525":11.00,"150576":49.10,"150385":49.10,
 "150551":24.20,"150388":25.90,"150016":19.70,"810620":50.30,"810366":47.10,
}

def price_chf(sku):
    p = ref.get(sku, {}).get("prix_catalogue", "")
    try:
        return float(p)
    except (TypeError, ValueError):
        return None

def eur(sku):
    return BLUPURA_EUR.get(ref.get(sku, {}).get("ref_fournisseur", ""))

# ---------- styles ----------
ARIAL   = "Arial"
H1      = Font(name=ARIAL, size=14, bold=True)
H2      = Font(name=ARIAL, size=11, bold=True)
BASE    = Font(name=ARIAL, size=10)
BOLD    = Font(name=ARIAL, size=10, bold=True)
ITAL    = Font(name=ARIAL, size=9, italic=True, color="595959")
WHITE_B = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
BLUE    = Font(name=ARIAL, size=10, color="0000FF")

HDR_FILL  = PatternFill("solid", fgColor="1F3864")
SUB_FILL  = PatternFill("solid", fgColor="D9E2F3")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
TOT_FILL  = PatternFill("solid", fgColor="E2EFDA")
thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)

MONEY = '#,##0.00;(#,##0.00);-'
INT   = '#,##0;(#,##0);-'

def header(ws, row, labels, widths=None):
    for i, lab in enumerate(labels, start=1):
        c = ws.cell(row=row, column=i, value=lab)
        c.font = WHITE_B; c.fill = HDR_FILL; c.border = BORD
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

def title(ws, text, sub=None):
    ws["A1"] = text; ws["A1"].font = H1
    if sub:
        ws["A2"] = sub; ws["A2"].font = ITAL

wb = Workbook()

# =====================================================================
# 1. SYNTHESE
# =====================================================================
ws = wb.active; ws.title = "Synthèse"
title(ws, "Commandes de pièces détachées par pays",
      "Extrait le 26.08.2026 — demandes Monday (board « Technical troubleshooting », BE WTR - Aftersales) "
      "croisées avec le référentiel SKU/prix (hub technicien + SharePoint).")
ws["A3"] = ("Deux demandes de commande explicites sont ouvertes dans Monday, toutes deux créées le 25.08.2026. "
            "Aucun autre pays n'a de demande d'achat enregistrée à ce jour.")
ws["A3"].font = BASE; ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("A3:G4"); ws.row_dimensions[3].height = 15; ws.row_dimensions[4].height = 15

header(ws, 6, ["Pays", "Produit principal", "Fournisseur", "Lignes",
               "Pièces (qté)", "Montant chiffrable (CHF)", "Lignes sans prix"],
       [16, 30, 16, 9, 13, 22, 16])

syn = [
    ("France",      "BOX 80 (BW-0001)",                      "Blupura",              14, 153, None,  10),
    ("Suisse",      "BOX 20 / BOX 15 / PRO1 / PRO2",         "Blupura + Borg&Overström", 5, 105, None,   3),
]
r = 7
for pays, prod, four, nl, qte, mnt, sans in syn:
    ws.cell(row=r, column=1, value=pays).font = BOLD
    ws.cell(row=r, column=2, value=prod).font = BASE
    ws.cell(row=r, column=3, value=four).font = BASE
    ws.cell(row=r, column=4, value=nl).font = BASE
    ws.cell(row=r, column=5, value=qte).font = BASE
    ws.cell(row=r, column=6, value=mnt).font = BASE
    ws.cell(row=r, column=7, value=sans).font = BASE
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = BORD
    ws.cell(row=r, column=5).number_format = INT
    ws.cell(row=r, column=6).number_format = MONEY
    r += 1

ws.cell(row=r, column=1, value="TOTAL").font = BOLD
ws.cell(row=r, column=4, value="=SUM(D7:D8)").font = BOLD
ws.cell(row=r, column=5, value="=SUM(E7:E8)").font = BOLD
ws.cell(row=r, column=6, value="=SUM(F7:F8)").font = BOLD
ws.cell(row=r, column=7, value="=SUM(G7:G8)").font = BOLD
for c in range(1, 8):
    ws.cell(row=r, column=c).border = BORD; ws.cell(row=r, column=c).fill = TOT_FILL
ws.cell(row=r, column=5).number_format = INT
ws.cell(row=r, column=6).number_format = MONEY

r += 3
ws.cell(row=r, column=1, value="Points bloquants avant passage des commandes").font = H2
r += 1
for txt in [
    "1. 13 des 19 lignes n'ont pas de prix catalogue : la base de septembre 2022 les marque « AS* » (prix à demander au service après-vente). "
    "Le montant chiffrable ci-dessus ne couvre donc que 5 lignes — un devis fournisseur est nécessaire pour le reste.",
    "2. La demande Suisse est rédigée en texte libre, sans SKU. 4 lignes sur 5 demandent une confirmation de référence "
    "(voir l'onglet « À clarifier ») et 2 lignes sont dupliquées dans Monday, ce qui rend la quantité ambiguë.",
    "3. Les pièces BOX 20 (BW-0987 → BW-1024) n'existent pas dans la base de prix de septembre 2022 : "
    "ni prix, ni référence fournisseur disponibles pour le ventilateur BOX 20 demandé par la Suisse.",
    "4. Plus largement, le référentiel ne permet pas de chiffrer grand-chose : sur les 294 références, "
    "70 seulement portent un prix exploitable (24 %). 138 sont marquées « AS* » et 86 sont absentes de la base de 2022. "
    "Reconstituer une base de prix à jour conditionne toute commande future, pas seulement celles-ci.",
]:
    c = ws.cell(row=r, column=1, value=txt); c.font = BASE
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=7)
    ws.row_dimensions[r].height = 15; ws.row_dimensions[r+1].height = 15
    r += 3

r += 1
ws.cell(row=r, column=1, value="Sources").font = H2
r += 1
for s in [
    "Demande France : Monday item 12889411743 « order spare parts france » — https://bewtr.monday.com/boards/2470912823/pulses/12889411743",
    "Demande Suisse : Monday item 12889827793 « spare parts suisse » — https://bewtr.monday.com/boards/2470912823/pulses/12889827793",
    "SKU & désignations : hub technicien (index.html, objet SPAREPARTS) — 294 références, 16 machines",
    "Prix catalogue CHF & réf. fournisseur : SharePoint « September 2022- BE WTR spare parts list.xlsm », onglet Database "
    "(/sites/Switzerland/Technical/05 Aftersales/Techs/02 Spare parts/), via docs/audit-spare-parts-2022.csv",
    "Prix d'achat Blupura EUR : SharePoint « Suggested Spare Parts list for Service - Price List 2022 - BE WTR.xlsx » (20 références)",
]:
    ws.cell(row=r, column=1, value=s).font = ITAL
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    r += 1

# =====================================================================
# 2. COMMANDE FRANCE
# =====================================================================
ws = wb.create_sheet("Commande France")
title(ws, "Commande France — BOX 80 (BW-0001)",
      "Source : Monday item 12889411743, créé le 25.08.2026 par Mathieu. Fournisseur unique : Blupura.")

header(ws, 4, ["SKU BW", "Désignation", "Réf. fournisseur", "Fournisseur",
               "Qté", "Prix cat. (CHF)", "Prix achat Blupura (EUR)",
               "Total CHF", "Total EUR", "Statut prix"],
       [11, 34, 15, 14, 7, 14, 20, 12, 12, 24])

FR = [("BW-0416",5),("BW-0183",20),("BW-0184",6),("BW-0417",3),("BW-0429",30),
      ("BW-0421",30),("BW-0419",2),("BW-0420",2),("BW-0428",10),("BW-0186",20),
      ("BW-0424",10),("BW-0425",5),("BW-0426",5),("BW-0195",5)]

r = 5
for sku, qty in FR:
    d = ref[sku]
    pc, pe = price_chf(sku), eur(sku)
    ws.cell(row=r, column=1, value=sku).font = BOLD
    ws.cell(row=r, column=2, value=d["designation"]).font = BASE
    ws.cell(row=r, column=3, value=d["ref_fournisseur"]).font = BASE
    ws.cell(row=r, column=4, value=d["fournisseur"]).font = BASE
    ws.cell(row=r, column=5, value=qty).font = BLUE
    ws.cell(row=r, column=6, value=pc).font = BASE
    ws.cell(row=r, column=7, value=pe).font = BASE
    ws.cell(row=r, column=8, value=f"=IF(N(F{r})=0,\"\",E{r}*F{r})").font = BASE
    ws.cell(row=r, column=9, value=f"=IF(N(G{r})=0,\"\",E{r}*G{r})").font = BASE
    if pc is None and pe is None:
        st, fill = "Prix à demander (AS*)", WARN_FILL
    elif pc is None:
        st, fill = "Prix achat Blupura seul", WARN_FILL
    elif pe is None:
        st, fill = "Prix catalogue seul", None
    else:
        st, fill = "Complet", None
    ws.cell(row=r, column=10, value=st).font = BASE
    for c in range(1, 11):
        cell = ws.cell(row=r, column=c); cell.border = BORD
        if fill and c == 10: cell.fill = fill
    for c in (6, 7, 8, 9):
        ws.cell(row=r, column=c).number_format = MONEY
    ws.cell(row=r, column=5).number_format = INT
    r += 1

ws.cell(row=r, column=1, value="TOTAL").font = BOLD
ws.cell(row=r, column=5, value=f"=SUM(E5:E{r-1})").font = BOLD
ws.cell(row=r, column=8, value=f"=SUM(H5:H{r-1})").font = BOLD
ws.cell(row=r, column=9, value=f"=SUM(I5:I{r-1})").font = BOLD
ws.cell(row=r, column=10, value=f'=COUNTIF(J5:J{r-1},"Prix à demander (AS*)")&" ligne(s) sans prix"').font = BOLD
for c in range(1, 11):
    cell = ws.cell(row=r, column=c); cell.border = BORD; cell.fill = TOT_FILL
ws.cell(row=r, column=5).number_format = INT
ws.cell(row=r, column=8).number_format = MONEY
ws.cell(row=r, column=9).number_format = MONEY
TOT_FR = r

r += 2
ws.cell(row=r, column=1,
        value="AS* = « Contact your BE WTR aftersales » dans la base de septembre 2022 : la pièce existe et porte une réf. fournisseur, "
              "mais aucun prix catalogue n'y est publié. Ces 10 lignes doivent faire l'objet d'un devis Blupura avant commande.").font = ITAL
ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=10)
r += 3
ws.cell(row=r, column=1,
        value="Le prix d'achat Blupura (EUR) provient de la liste fournisseur 2022 (20 références seulement). Le prix catalogue (CHF) est un prix BE WTR, "
              "pas un prix d'achat : sur les 4 lignes où les deux existent, l'écart va de 1,8x à 2,0x.").font = ITAL
ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=10)
r += 3
ws.cell(row=r, column=1, value="Bleu = valeur saisie depuis Monday. Noir = calcul ou donnée du référentiel.").font = ITAL

# =====================================================================
# 3. COMMANDE SUISSE
# =====================================================================
ws = wb.create_sheet("Commande Suisse")
title(ws, "Commande Suisse — multi-produits",
      "Source : Monday item 12889827793, créé le 25.08.2026. Demande en texte libre, sans SKU : les références ci-dessous sont des propositions à valider.")

header(ws, 4, ["Demande Monday (verbatim)", "Produit", "SKU proposé", "Désignation référentiel",
               "Réf. fourn.", "Fournisseur", "Qté", "Prix cat. (CHF)", "Total CHF", "Fiabilité du mapping"],
       [40, 18, 12, 32, 12, 18, 7, 14, 12, 26])

CH = [
    ("x30 Ventilateur BOX 20 (compatible aussi BOX 15)", "BOX 20 / BOX 15", "BW-0988", 30,
     "À trancher — voir « À clarifier »"),
    ("x20 Transformateurs 230-12V pour PRO2 ancienne génération", "PRO2 V1", "BW-0301", 20,
     "À confirmer — tension différente"),
    ("15x Ramasse-goutte PRO2 V1", "PRO2 V1", "BW-0507", 15,
     "Probable — ligne dupliquée dans Monday"),
    ("20x Carte électronique PRO1 (Pin)", "PRO1", "BW-0159", 20,
     "À confirmer — 2 cartes PRO1 possibles"),
    ("20x Sonde BOX PRO2 V2", "PRO2 V2", "", 20,
     "Non résolu — aucune sonde PRO2 au référentiel"),
]

r = 5
for verbatim, prod, sku, qty, fiab in CH:
    d = ref.get(sku, {})
    pc = price_chf(sku) if sku else None
    ws.cell(row=r, column=1, value=verbatim).font = BLUE
    ws.cell(row=r, column=2, value=prod).font = BASE
    ws.cell(row=r, column=3, value=sku or "—").font = BOLD
    ws.cell(row=r, column=4, value=d.get("designation") or "—").font = BASE
    ws.cell(row=r, column=5, value=d.get("ref_fournisseur") or "—").font = BASE
    ws.cell(row=r, column=6, value=d.get("fournisseur") or "—").font = BASE
    ws.cell(row=r, column=7, value=qty).font = BLUE
    ws.cell(row=r, column=8, value=pc).font = BASE
    ws.cell(row=r, column=9, value=f"=IF(N(H{r})=0,\"\",G{r}*H{r})").font = BASE
    ws.cell(row=r, column=10, value=fiab).font = BASE
    for c in range(1, 11):
        cell = ws.cell(row=r, column=c); cell.border = BORD
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if c == 10: cell.fill = WARN_FILL
    ws.cell(row=r, column=7).number_format = INT
    ws.cell(row=r, column=8).number_format = MONEY
    ws.cell(row=r, column=9).number_format = MONEY
    ws.row_dimensions[r].height = 30
    r += 1

ws.cell(row=r, column=1, value="TOTAL").font = BOLD
ws.cell(row=r, column=7, value=f"=SUM(G5:G{r-1})").font = BOLD
ws.cell(row=r, column=9, value=f"=SUM(I5:I{r-1})").font = BOLD
for c in range(1, 11):
    cell = ws.cell(row=r, column=c); cell.border = BORD; cell.fill = TOT_FILL
ws.cell(row=r, column=7).number_format = INT
ws.cell(row=r, column=9).number_format = MONEY
TOT_CH = r

r += 2
ws.cell(row=r, column=1,
        value="Quantités retenues : la demande Monday répète deux lignes à l'identique (ramasse-goutte 15x, carte PRO1 20x). "
              "La lecture retenue ici est une duplication de saisie — donc 15 et 20, pas 30 et 40. À confirmer avec le demandeur avant commande.").font = ITAL
ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=10)
r += 3
ws.cell(row=r, column=1,
        value="Aucun total fiable n'est calculable en l'état : 4 lignes sur 5 attendent une validation de référence. "
              "Le total CHF ci-dessus ne couvre que les SKU proposés disposant d'un prix catalogue.").font = ITAL
ws.merge_cells(start_row=r, start_column=1, end_row=r+1, end_column=10)

# =====================================================================
# 4. A CLARIFIER
# =====================================================================
ws = wb.create_sheet("À clarifier")
title(ws, "Points à trancher avant de passer commande",
      "Chaque ligne bloque tout ou partie d'une commande.")

header(ws, 4, ["#", "Pays", "Sujet", "Constat", "Options / action", "Impact"],
       [5, 12, 30, 55, 55, 26])

CLAR = [
 (1, "France", "10 lignes sans prix (AS*)",
  "BW-0416, BW-0417, BW-0429, BW-0421, BW-0419, BW-0420, BW-0428, BW-0424, BW-0425, BW-0426 portent la mention « AS* » "
  "dans la base de septembre 2022 : réf. fournisseur connue, prix non publié.",
  "Demander un devis à Blupura sur ces 10 réf. fournisseur (110264, 130048A, 750019, 140118, 110138, 110137, 150052, 140099, 150094, 150072).",
  "Commande France non chiffrable à 100 %"),
 (2, "France", "Prix catalogue ≠ prix d'achat",
  "Sur les 4 lignes où les deux prix existent, le catalogue CHF vaut environ le double du prix d'achat Blupura EUR "
  "(ex. BW-0195 : 40 CHF catalogue vs 21,60 EUR achat).",
  "Utiliser les prix d'achat Blupura pour le budget de commande, pas le catalogue.",
  "Budget surestimé d'environ 2x si le catalogue est utilisé"),
 (3, "Suisse", "Ventilateur BOX 20 : quel SKU ?",
  "La demande dit « Ventilateur BOX 20, compatible aussi BOX 15, à voir quel fournisseur est le moins cher ». "
  "Deux références coexistent : BW-0988 (Axial fan 120x120x25, BOX 20 — absente de la base prix, aucun fournisseur) "
  "et BW-0191 (Fan 120x120, BOX 15 — Blupura 120232, 42 CHF). La liste Blupura cote aussi un « Fan 120x120 » réf. 750244 à 22,50 EUR.",
  "Confirmer que le ventilateur BOX 20 et le ventilateur BOX 15 sont bien interchangeables, puis arbitrer entre Blupura 120232 et 750244.",
  "30 pièces — ligne la plus volumineuse de la demande suisse"),
 (4, "Suisse", "Transformateur 230-12V PRO2 V1",
  "Aucun transformateur 230-12V au référentiel. Les candidats sont en 24V : BW-0301 (PRO2 Power adapter 230V/24VDC, B&O 174391, 40 CHF) "
  "et BW-0293 (BAR2 Transformator, Blupura 150016, 39 CHF).",
  "Faire préciser la tension et le modèle PRO2 concerné par le technicien, ou relever la référence sur une pièce en stock.",
  "20 pièces — risque d'erreur de tension"),
 (5, "Suisse", "Carte électronique PRO1 : deux candidats",
  "BW-0159 « PRO1 - PCB Control Board » (B&O 701052, 72,50 CHF) et BW-0164 « PRO1 - PCB Control board - A » (B&O 637107, 195 CHF). "
  "La mention « (Pin) » de la demande ne tranche pas.",
  "Faire préciser la génération PRO1. Écart de prix : 122,50 CHF/pièce, soit 2 450 CHF sur 20 pièces.",
  "20 pièces — écart de 2 450 CHF"),
 (6, "Suisse", "Sonde BOX PRO2 V2 introuvable",
  "Le référentiel ne contient aucune sonde pour PRO2. Les sondes existantes sont PRO1 (BW-0646 réfrigération, BW-0647 niveau) "
  "et BOX 30 (BW-1068 niveau carbo, BW-1069 température).",
  "Faire identifier la pièce par le technicien (photo ou réf. B&O), puis créer le SKU si elle est absente du référentiel.",
  "20 pièces — ligne non commandable"),
 (7, "Suisse", "Deux lignes dupliquées dans Monday",
  "« 15x Ramasse goutte PRO2 V1 » et « 20x Carte électronique PRO1 (Pin) » apparaissent chacune deux fois à l'identique dans la demande.",
  "Confirmer s'il s'agit d'une duplication de saisie (15 et 20) ou de deux besoins distincts (30 et 40).",
  "35 pièces d'écart possible"),
 (8, "Global", "Base de prix vieille de 4 ans",
  "La seule base de prix complète est celle de septembre 2022. Les machines BOX 20, BOX 80 I et BOX 120 I "
  "(82 SKU, BW-0987 → BW-1072) n'y figurent pas du tout.",
  "Demander des tarifs à jour à Blupura et Borg&Overström, et compléter la base pour les 3 machines manquantes.",
  "Tous les pays"),
]

r = 5
for n, pays, sujet, constat, action, impact in CLAR:
    ws.cell(row=r, column=1, value=n).font = BOLD
    ws.cell(row=r, column=2, value=pays).font = BOLD
    ws.cell(row=r, column=3, value=sujet).font = BASE
    ws.cell(row=r, column=4, value=constat).font = BASE
    ws.cell(row=r, column=5, value=action).font = BASE
    ws.cell(row=r, column=6, value=impact).font = BASE
    for c in range(1, 7):
        cell = ws.cell(row=r, column=c); cell.border = BORD
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 58
    r += 1

# =====================================================================
# 5. CLAIMS OUVERTS
# =====================================================================
ws = wb.create_sheet("Claims ouverts")
title(ws, "Claims ouverts par pays et produit — besoin de pièces à venir",
      "Board Monday « Technical troubleshooting », statuts non clôturés au 26.08.2026. Indicateur de demande, hors commandes formalisées.")

header(ws, 4, ["Pays", "Produit", "Nouveau (à réparer)", "En cours", "Chez le fournisseur", "Bloqué", "Total ouvert"],
       [22, 34, 18, 12, 20, 11, 14])

CLAIMS = [
 ("France",  "BW-0001 BOX 80",                    9, 0, 1, 0),
 ("Suisse",  "BW-0074 BOX 30",                    9, 0, 0, 0),
 ("Suisse",  "BW-0001 BOX 80",                    7, 0, 0, 0),
 ("Suisse",  "BW-0271 PRO2 White",                5, 0, 0, 0),
 ("Suisse",  "BW-0272.01 PRO2 Black",             3, 1, 0, 0),
 ("Suisse",  "BW-0596 BOX 20 Home",               2, 2, 1, 0),
 ("Suisse",  "BW-0067 BOX30 B",                   2, 0, 0, 0),
 ("Suisse",  "BW-0272 PRO2 Black",                1, 0, 0, 0),
 ("Suisse",  "BW-0271.01 PRO2 Silver",            1, 0, 0, 0),
 ("Suisse",  "BW-0136 BAR2 double portion ctrl",  1, 0, 0, 0),
 ("Suisse",  "BW-0042 AQTiV COMBI",               1, 0, 0, 0),
 ("Suisse",  "BW-0045 AQTiV COMBI H",             1, 0, 0, 0),
 ("Suisse",  "BW-0617 BE CONNECT",                1, 0, 0, 0),
 ("UK",      "BW-0136 BAR2 double portion ctrl",  1, 0, 0, 0),
 ("Émirats", "BW-0074 BOX 30",                    1, 0, 0, 0),
 ("Émirats", "Tower",                             1, 0, 0, 0),
 ("Espagne", "(produit non renseigné)",           1, 0, 0, 0),
 ("Canada",  "BW-0001 BOX 80",                    0, 0, 0, 1),
]

r = 5
for pays, prod, nw, ip, sup, stk in CLAIMS:
    ws.cell(row=r, column=1, value=pays).font = BOLD
    ws.cell(row=r, column=2, value=prod).font = BASE
    for i, v in enumerate((nw, ip, sup, stk), start=3):
        c = ws.cell(row=r, column=i, value=v); c.font = BASE; c.number_format = INT
    ws.cell(row=r, column=7, value=f"=SUM(C{r}:F{r})").font = BOLD
    ws.cell(row=r, column=7).number_format = INT
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = BORD
    r += 1

ws.cell(row=r, column=1, value="TOTAL").font = BOLD
for col in range(3, 8):
    L = get_column_letter(col)
    ws.cell(row=r, column=col, value=f"=SUM({L}5:{L}{r-1})").font = BOLD
    ws.cell(row=r, column=col).number_format = INT
for c in range(1, 8):
    cell = ws.cell(row=r, column=c); cell.border = BORD; cell.fill = TOT_FILL

r += 2
ws.cell(row=r, column=1,
        value="Les deux demandes de commande (France BOX 80, Suisse PRO2 Black) sont elles-mêmes enregistrées comme claims « New (to repair) » "
              "et sont incluses dans les compteurs ci-dessus.").font = ITAL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
r += 2
ws.cell(row=r, column=1,
        value="La Suisse concentre 38 des 53 claims ouverts. La France en a 10, tous sur BOX 80 — cohérent avec la commande passée.").font = ITAL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)

# =====================================================================
# 6. REFERENTIEL
# =====================================================================
ws = wb.create_sheet("Référentiel SKU & prix")
title(ws, "Référentiel pièces détachées — 294 SKU",
      "Hub technicien (désignations, machines) croisé avec la base SharePoint de septembre 2022 (fournisseur, réf., prix CHF) "
      "et la liste d'achat Blupura 2022 (prix EUR).")

header(ws, 4, ["SKU BW", "Désignation", "Machines (hub)", "Fournisseur", "Réf. fournisseur",
               "Prix cat. (CHF)", "Prix achat Blupura (EUR)", "Remarque"],
       [11, 40, 34, 18, 15, 14, 20, 26])

r = 5
for d in sorted(rows_ref, key=lambda x: x["ref"]):
    sku = d["ref"]
    pc, pe = price_chf(sku), eur(sku)
    raw = d["prix_catalogue"]
    rem = ""
    if raw == "AS*":
        rem = "Prix à demander (AS*)"
    elif raw == "" and not d["fournisseur"]:
        rem = "Absente de la base 2022"
    ws.cell(row=r, column=1, value=sku).font = BOLD
    ws.cell(row=r, column=2, value=d["designation"]).font = BASE
    ws.cell(row=r, column=3, value=d["machines_hub"]).font = BASE
    ws.cell(row=r, column=4, value=d["fournisseur"]).font = BASE
    ws.cell(row=r, column=5, value=d["ref_fournisseur"]).font = BASE
    ws.cell(row=r, column=6, value=pc).font = BASE
    ws.cell(row=r, column=7, value=pe).font = BASE
    ws.cell(row=r, column=8, value=rem).font = BASE
    ws.cell(row=r, column=6).number_format = MONEY
    ws.cell(row=r, column=7).number_format = MONEY
    for c in range(1, 9):
        ws.cell(row=r, column=c).border = BORD
    if rem:
        ws.cell(row=r, column=8).fill = WARN_FILL
    r += 1

ws.freeze_panes = "A5"
ws.auto_filter.ref = f"A4:H{r-1}"

for s in wb.worksheets:
    if s.title != "Référentiel SKU & prix":
        s.sheet_view.showGridLines = False

ws_syn = wb["Synthèse"]
for cell, formula in (("F7", f"='Commande France'!H{TOT_FR}"),
                      ("F8", f"='Commande Suisse'!I{TOT_CH}")):
    ws_syn[cell] = formula
    ws_syn[cell].font = BASE
    ws_syn[cell].number_format = MONEY
    ws_syn[cell].border = BORD

wb.save(OUT)
print("OK", OUT)
